import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException
from directory_helper import move_file, create_directory, is_file

def _get_sheets(wb):
    return wb.worksheets

class ExcelManager:
    _base_path = None
    _result_file_name = 'consolidation.xlsx'
    processed_folder_name = 'Processed'
    not_processed_folder_name = 'Not applicable'
    result_folder_name = 'Result'
    folders = [processed_folder_name, not_processed_folder_name, result_folder_name]

    def _get_absolute_path(self, concat=''):
        return f'{self._base_path}/{concat}'

    def _ignore_file(self, file):
        move_file(self._get_absolute_path(file), f'{self.not_processed_path}/{file}')

    def _accept_file(self, file):
        move_file(self._get_absolute_path(file), f'{self.processed_path}/{file}')

    def _create_output_directory(self):
        create_directory(self.output_path)

    @property
    def output_path(self):
        return self._get_absolute_path(self.result_folder_name)

    @property
    def output_file_path(self):
        return f'{self.output_path}/{self._result_file_name}'

    @property
    def not_processed_path(self):
        return self._get_absolute_path(self.not_processed_folder_name)

    @property
    def processed_path(self):
        return self._get_absolute_path(self.processed_folder_name)

    def set_base_path(self, path):
        self._base_path = path

    def get_output_file(self):
        created = False
        self._create_output_directory()

        if not is_file(self.output_file_path):
            create_directory(self.output_file_path)
            wb = Workbook()
            wb.save(self.output_file_path)
            created = True

        return load_workbook(self.output_file_path), created

    def sheets_copy(self, wb):
        output, created = self.get_output_file()
        if created:
            output.remove(output.worksheets[0])

        for sheet in _get_sheets(wb):
            output_sheet = output.create_sheet(title=sheet.title)
            for row_data in sheet.values:
                output_sheet.append(row_data)

        output.save(self.output_file_path)

    def process(self, file):
        if not self._base_path:
            raise 'base path is required'
        if not is_file(self._get_absolute_path(file)):
            self._ignore_file(file)
            print('not is a file:', file)
            return

        try:
            wb = load_workbook(self._get_absolute_path(file))
        except InvalidFileException:
            self._ignore_file(file)
            print('invalid file:', file)
            return

        self.sheets_copy(wb)
        self._accept_file(file)
